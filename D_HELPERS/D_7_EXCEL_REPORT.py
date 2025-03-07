from openpyxl import Workbook
from D_HELPERS.D_8_EXCEL_CHART import ExcelChartGenerator  

class ExcelReport:
    """ Class for generating Excel reports. """

    @staticmethod
    def create_excel_report(excel_file, sorted_clues, sorted_categories, sorted_answers):
        """ Generate an Excel report with analysis results. """
        
        wb = Workbook()

        # 📌 SHEET 1: CLUE RESULTS
        ws_clue = wb.active
        ws_clue.title = "Clue Results"
        ws_clue.append(["Clue", "Category", "Answer", "Occurrences"])

        for (clue, category, answer), count in sorted_clues:
            ws_clue.append([clue, category, answer, count])

        ExcelChartGenerator.add_chart(ws_clue, "Occurrences", "Clue Results")

        # 📌 SHEET 2: CATEGORIES
        ws_category = wb.create_sheet(title="Categories")
        ws_category.append(["Category", "Occurrences"])

        for category, count in sorted_categories:
            ws_category.append([category, count])

        ExcelChartGenerator.add_chart(ws_category, "Occurrences", "Category Results")

        # 📌 SHEET 3: ANSWERS
        ws_answer = wb.create_sheet(title="Answers")
        ws_answer.append(["Answer", "Occurrences"])

        for answer, count in sorted_answers:
            ws_answer.append([answer, count])

        ExcelChartGenerator.add_chart(ws_answer, "Occurrences", "Answer Results")

        # 📌 Save the Excel file
        wb.save(excel_file)
        print(f"\n✅ Excel report successfully saved at: {excel_file}")
